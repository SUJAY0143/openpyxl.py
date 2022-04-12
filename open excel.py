from time import sleep
from selenium import webdriver
import openpyxl
from time import sleep

driver = webdriver.Chrome(executable_path="c:\\Users\\sujay\\chromedriver_win32\\chromedriver.exe")
#driver.get('https://www.google.com/')
sleep(2)

book = openpyxl.load_workbook("D:\SELENIUM\pyxldemo.xlsx")
sheet = book.active
#cell = sheet.cell(row=1, column=2)
#print(cell.value)

#to print all the values from the sheet as matrix
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=2,column=1).value=='test1':
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value,end=" ")
        print()

driver.close()